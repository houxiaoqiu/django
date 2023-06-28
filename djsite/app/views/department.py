from django.shortcuts import redirect, render,HttpResponse
from app import models
import os
from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl import load_workbook

# 部门列表
def department_list(request):
    queryset = models.Department.objects.all()
    return render(request,'department_list.html',{'queryset': queryset})

# 部门新增
def department_add(request):
    if request.method == 'GET':
        return render(request,'department_add.html')
    else:
        title = request.POST.get('title')
        models.Department.objects.create(title=title)
        return redirect("/department/list/")

# 部门删除
def department_delete(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/department/list/")

# 部门编辑
def department_edit(request,nid):
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request,'department_edit.html',{"row_object":row_object})
    else:
        title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect("/department/list/")
    
# 部门上传文件
def department_multi(request):
    ## 获取上传excel文件对象
    file_object = request.FILES.get("xlsx")
    ## 获取取excel文件工作簿对象
    wb = load_workbook(file_object)
    ws = wb.worksheets[0]
    ## 循环获取每行数据,并将数据导入数据库Department表中
    for row in ws.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/department/list/')

